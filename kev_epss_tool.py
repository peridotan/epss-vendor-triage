# kev_epss_tool.py
# -*- coding: utf-8 -*-

import argparse
import csv
import gzip
import io
from datetime import datetime
from pathlib import Path

import requests


KEV_URL = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
EPSS_URL = "https://epss.cyentia.com/epss_scores-current.csv.gz"


def fetch_json(url, timeout=30):
    response = requests.get(url, timeout=timeout)
    response.raise_for_status()
    return response.json()

def fetch_epss(timeout=30):
    response = requests.get(EPSS_URL, timeout=timeout)
    response.raise_for_status()

    epss_map = {}

    with gzip.GzipFile(fileobj=io.BytesIO(response.content)) as gz:
        text = gz.read().decode("utf-8")

        # EPSS CSVには先頭に # で始まるメタ情報行が入るため除外する
        lines = [
            line for line in text.splitlines()
            if line.strip() and not line.startswith("#")
        ]

        reader = csv.DictReader(lines)

        for row in reader:
            cve = row.get("cve")
            if not cve:
                continue

            epss_map[cve] = {
                "epss_score": float(row.get("epss", 0) or 0),
                "epss_percentile": float(row.get("percentile", 0) or 0),
            }

    print(f"EPSS records loaded: {len(epss_map)}")
    return epss_map

def build_rows(kev_data, epss_map, epss_threshold=0.0, percentile_threshold=0.0):
    rows = []

    for item in kev_data.get("vulnerabilities", []):
        cve = item.get("cveID", "")

        epss = epss_map.get(cve, {})
        epss_score = epss.get("epss_score")
        epss_percentile = epss.get("epss_percentile")

        if epss_score is None:
            continue

        if epss_score < epss_threshold:
            continue

        if epss_percentile < percentile_threshold:
            continue

        rows.append({
            "cve": cve,
            "vendor": item.get("vendorProject", ""),
            "product": item.get("product", ""),
            "vulnerability_name": item.get("vulnerabilityName", ""),
            "description": item.get("shortDescription", ""),
            "date_added": item.get("dateAdded", ""),
            "due_date": item.get("dueDate", ""),
            "known_ransomware_campaign_use": item.get("knownRansomwareCampaignUse", ""),
            "required_action": item.get("requiredAction", ""),
            "notes": item.get("notes", ""),
            "epss_score": epss_score,
            "epss_percentile": epss_percentile,
        })

    rows.sort(key=lambda r: r["epss_score"], reverse=True)
    return rows


def write_csv(rows, output_path):
    fieldnames = [
        "cve",
        "vendor",
        "product",
        "vulnerability_name",
        "description",
        "date_added",
        "due_date",
        "known_ransomware_campaign_use",
        "required_action",
        "notes",
        "epss_score",
        "epss_percentile",
    ]

    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "KEV_EPSS"

    fieldnames = [
        "cve",
        "vendor",
        "product",
        "vulnerability_name",
        "description",
        "date_added",
        "due_date",
        "known_ransomware_campaign_use",
        "required_action",
        "notes",
        "epss_score",
        "epss_percentile",
    ]

    ws.append(fieldnames)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(fieldnames, start=1):
        max_len = len(col_name)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Fetch CISA KEV and EPSS data, then export enriched vulnerability data.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="""
Examples:
  python kev_epss_tool.py --xlsx
  python kev_epss_tool.py --epss-threshold 0.7 --xlsx
  python kev_epss_tool.py --percentile-threshold 0.99 --xlsx
  python kev_epss_tool.py --epss-threshold 0.7 --percentile-threshold 0.99 --xlsx
  python kev_epss_tool.py -o output/kev_epss_result.csv --xlsx

Typical workflow:
  1. Fetch KEV + EPSS:
     python kev_epss_tool.py --xlsx

  2. Analyze vendors:
     python vendor_from_epss_csv.py kev_epss_result.csv --score 0.9
"""
    )

    parser.add_argument(
        "--epss-threshold",
        type=float,
        default=0.0,
        help="Minimum EPSS score threshold. Example: 0.7"
    )

    parser.add_argument(
        "--percentile-threshold",
        type=float,
        default=0.0,
        help="Minimum EPSS percentile threshold. Example: 0.99"
    )

    parser.add_argument(
        "--output",
        "-o",
        default="kev_epss_result.csv",
        help="Output CSV file path. Default: kev_epss_result.csv"
    )

    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Also export XLSX file."
    )

    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="HTTP request timeout in seconds. Default: 30"
    )

    # Show help when no arguments are provided
    import sys
    if len(sys.argv) == 1:
        parser.print_help()
        return

    args = parser.parse_args()

    print("Fetching CISA KEV...")
    kev_data = fetch_json(KEV_URL, timeout=args.timeout)

    print("Fetching EPSS...")
    epss_map = fetch_epss(timeout=args.timeout)

    print("Building result...")
    rows = build_rows(
        kev_data,
        epss_map,
        epss_threshold=args.epss_threshold,
        percentile_threshold=args.percentile_threshold,
    )

    output_csv = Path(args.output)
    write_csv(rows, output_csv)

    print(f"CSV written: {output_csv}")
    print(f"Total rows : {len(rows)}")

    if args.xlsx:
        xlsx_path = output_csv.with_suffix(".xlsx")
        write_xlsx(rows, xlsx_path)
        print(f"XLSX written: {xlsx_path}")


if __name__ == "__main__":
    main()
